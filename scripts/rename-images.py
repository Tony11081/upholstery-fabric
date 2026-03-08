import os
from openpyxl import load_workbook

# Load Excel
excel_path = r'C:\Users\Administrator\Desktop\200个.xlsx'
output_dir = r'e:\购物网站\luxury-shop\public\products'

wb = load_workbook(excel_path)
ws = wb.active

# Build row to image mapping
images = ws._images
row_to_images = {}

for i, img in enumerate(images):
    row = 0
    if hasattr(img.anchor, '_from'):
        row = img.anchor._from.row
    elif hasattr(img.anchor, 'row'):
        row = img.anchor.row

    if row not in row_to_images:
        row_to_images[row] = []
    row_to_images[row].append(i + 1)

# Read product data and save with proper names
print("Row -> Image mapping:")
for row_num in range(2, ws.max_row + 1):
    product_id = ws.cell(row=row_num, column=5).value  # Column E = 商品Id
    desc = ws.cell(row=row_num, column=2).value  # Column B = 商品名称/描述

    if product_id and desc:
        # Row in images is 0-indexed, Excel row is 1-indexed
        img_row = row_num - 1
        if img_row in row_to_images:
            img_indices = row_to_images[img_row]
            # Rename first image to product ID
            old_name = f"product_{img_indices[0]}.png"
            new_name = f"{product_id}.png"
            old_path = os.path.join(output_dir, old_name)
            new_path = os.path.join(output_dir, new_name)
            if os.path.exists(old_path):
                os.rename(old_path, new_path)
                print(f"Row {row_num}: {old_name} -> {new_name}")

print("\nDone!")
